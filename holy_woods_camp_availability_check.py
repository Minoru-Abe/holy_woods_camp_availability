import pyautogui
import webbrowser
import openpyxl
from time import sleep
import subprocess
import datetime
import line_util
import os

cwd = os.getcwd()
CHAR_MONTH = "月"
CHAR_TREE = "ツリー"
header_column = 1
LINECODE = "\n"
COMMA = ","
CHAR_NOT_AVAILABLE = "✕"
CHAR_HOLIDAY = "休"
CHAR_AVAILABLE = "Available"
CHAR_UNAVAILABLE = "Unavailable"
LINE_MAX_NO = 16
OUTPUT_FILE_NAME = "holy_woods_camp_availability.xlsx"
EXCEL = cwd + "/" + OUTPUT_FILE_NAME
EXCEL_SHEET = "availability"

#open holy woods availability homepage
holy_woods_url_file = open("param_holy_woods_url.csv", "r")
file_header = next(holy_woods_url_file)
url = holy_woods_url_file.readline().replace(LINECODE,"")
webbrowser.open(url)

#Sleep to wait for the homepage opening
sleep(3)

#Click to enable the page
pyautogui.click(100,100)

#Copy the availability homepage by pressing CNTL+A & CNTL+C
pyautogui.hotkey('ctrl', 'a')
pyautogui.hotkey('ctrl', 'c')

#Create output excel file
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = EXCEL_SHEET
wb.save(OUTPUT_FILE_NAME)

#Open the created excel file with subprocess
subprocess.Popen(["start",EXCEL], shell=True)

#Sleep to wait for the excel file opening
sleep(2)

#Paste copied availability information to the excel and close the excel file
pyautogui.hotkey('ctrl', 'v')
sleep(1)
pyautogui.hotkey('ctrl', 's')
sleep(1)
pyautogui.hotkey('alt', 'f4')

#Open the created excel file with openpyxl
wb = openpyxl.load_workbook(OUTPUT_FILE_NAME)
sheet = wb[EXCEL_SHEET]

#Make a list to check what row is related to tree house availability.
#The list is two dimensional list as followings
#[month, row_number_of_treehouse_availability]
#For example like the following
#[[10, 20], [11, 40], [12, 60]]
#The above example shows tree house availability in October is written in row number 20 in the excel, 40 for November, 60 for December
tree_house_row_list = []
month = 0
tree_row = 0
for rownum in range(1, 150):
    cellvalue = sheet.cell(column=header_column, row=rownum).value
    if cellvalue is None:
        continue
    if CHAR_MONTH in cellvalue:
        month = cellvalue.replace(LINECODE,"").replace(CHAR_MONTH,"")
    if CHAR_TREE in cellvalue:
        tree_row_number = rownum
        list = [month, tree_row_number]
        tree_house_row_list.append(list)

#Create availability table
#This part will make a four dimensional list like the following
#[["2022", "10", "01", ""], ["2022", "10", "02", "✕"], .....]
#Each contens includes year, month, day and availability on holy woods homepage
#Regarding the above example, tree house is available on 2022/10/01, not available on 2022/10/02, something like that.
#This list also includes some wrong date like 2022/02/31 since this part will simply search the excel file from column 2 to 32 for each row
current_year = datetime.date.today().year
row_count = 0
result_list = []
for tree_house in tree_house_row_list:
    row_count = row_count + 1
    availability_month = format(int(tree_house[0]),"02")
    target_row = tree_house[1]
    for day in range(1, 32):
        availability = sheet.cell(column=day+1, row=target_row).value
        #Set availability based on the character like ✕ or 休 on the homepage
        if availability is None:
            availability = CHAR_AVAILABLE
        elif CHAR_NOT_AVAILABLE in availability or CHAR_HOLIDAY in availability:
            availability = CHAR_UNAVAILABLE
        else:
            availability = CHAR_AVAILABLE
        #Convert year if the month is January or February. It should be next year from current year in some cases.
        if row_count == 2 and availability_month == "01":
            availability_year = str(current_year + 1)
        elif row_count == 3 and (availability_month == "01" or availability_month == "02"):
            availability_year = str(current_year + 1)
        else:
            availability_year = str(current_year)
        availability_day = format(day,"02")
        result_list.append([availability_year, availability_month, availability_day, availability])
print(result_list)

sent_list = []
#Check whether the date is Saturday or not, and if Saturday send out notification
#If the date is the past we don't send out notification

#Get current date with yyyymmdd format
current_date = "{0:%Y%m%d}".format(datetime.datetime.now())

for result in result_list:
    target_date = result[0] + result[1] + result[2]
    availability = result[3]
    if int(current_date) <= int(target_date):
        try:
            weekday = datetime.datetime(int(result[0]), int(result[1]), int(result[2])).weekday()
            #Return Saturday availability based on weekday
            if weekday == 5:
                result_line = target_date + COMMA + availability
                sent_list.append(result_line)
        except ValueError:
            continue
print(sent_list)

#Send the result message to line
message_counter = 0
message_sender = line_util.SendNotification
messages_to_be_sent = ""
for message in sent_list:
    messages_to_be_sent = messages_to_be_sent + message + LINECODE
    message_counter += 1
    if message_counter == LINE_MAX_NO:
        message_sender.send_message(messages_to_be_sent)
        message_counter = 0
        messages_to_be_sent = ""
#For the case like there are only three messages, or there are 7 messages (not a multiple of 4)
if message_counter > 0:
    message_sender.send_message(messages_to_be_sent)
